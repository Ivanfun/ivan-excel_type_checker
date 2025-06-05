from fastapi import FastAPI, UploadFile, File, Request, HTTPException, status
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
import pandas as pd
import shutil, os, tempfile
import openpyxl # Explicit import for openpyxl.Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font # 導入 Font
from openpyxl.utils import quote_sheetname, get_column_letter
import logging

app = FastAPI()
templates = Jinja2Templates(directory="frontend")

# 設定日誌記錄
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 使用更安全的臨時目錄管理方式
if os.path.exists("temp_output"):
    try:
        shutil.rmtree("temp_output")
        logging.info("已清空舊的 'temp_output' 目錄。")
    except OSError as e:
        logging.error(f"清理舊的 'temp_output' 目錄失敗: {e}")
os.makedirs("temp_output", exist_ok=True)
TEMP_OUTPUT_DIR = "temp_output"
logging.info(f"臨時輸出目錄已建立: {TEMP_OUTPUT_DIR}")


# 設定最大檔案大小限制為 10 MB (可根據需求調整)
MAX_FILE_SIZE_MB = 10

@app.get("/", response_class=HTMLResponse)
async def main(request: Request):
    """
    根目錄路由，顯示上傳表單頁面。
    """
    logging.info("接收到 GET / 請求，返回主頁面。")
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/upload/")
async def upload_file(request: Request, file: UploadFile = File(...)):
    """
    處理檔案上傳和分析。
    """
    input_path = None
    temp_input_dir = None

    try:
        logging.info(f"接收到檔案上傳請求：{file.filename} (MIME: {file.content_type})")

        # 1. 檔案大小檢查
        file.file.seek(0, os.SEEK_END)
        file_size = file.file.tell()
        file.file.seek(0)

        if file_size > MAX_FILE_SIZE_MB * 1024 * 1024:
            logging.warning(f"檔案 {file.filename} (大小: {file_size} bytes) 超出大小限制。")
            return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={
                "success": False,
                "message": f"上傳檔案大小超出限制。檔案大小不能超過 {MAX_FILE_SIZE_MB} MB。"
            })

        # 2. 儲存上傳的檔案到臨時目錄
        temp_input_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_input_dir, file.filename)
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        logging.info(f"檔案已暫存到: {input_path}")

        file_base_name, file_ext = os.path.splitext(file.filename)
        file_ext = file_ext.lower()

        supported_extensions = [".xlsx", ".xls", ".xlsm"]
        if file_ext not in supported_extensions:
            logging.warning(f"不支援的檔案格式: {file.filename} ({file_ext})")
            return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={
                "success": False,
                "message": f"不支援的檔案格式。僅支援 {', '.join(supported_extensions)} 檔案。"
            })
        
        # 3. 使用 pandas 讀取 Excel 檔案
        try:
            df = pd.read_excel(input_path, sheet_name=0)
        except Exception as e:
            try:
                xl = pd.ExcelFile(input_path)
                if len(xl.sheet_names) == 1:
                    df = pd.read_excel(input_path, sheet_name=xl.sheet_names[0])
                else:
                    logging.error(f"讀取 Excel 檔案 {file.filename} 時發生錯誤: {e}. 多個工作表且未指定。")
                    raise ValueError(f"Excel 檔案包含多個工作表，請確保僅上傳包含一個工作表的檔案，或修改程式以處理特定工作表。")
            except Exception as e_inner:
                logging.error(f"讀取 Excel 檔案 {file.filename} 時發生錯誤: {e_inner}。請確認已安裝處理 '{file_ext}' 格式所需的引擎 (如 xlrd for .xls)。")
                raise ValueError(f"無法讀取 Excel 檔案，請確認檔案格式正確、包含有效數據，且已安裝對應處理引擎。詳情: {e_inner}")


        # 4. 驗證必要欄位
        required_columns = {"Name", "Data Type"}
        if not required_columns.issubset(df.columns):
            missing_columns = required_columns - set(df.columns)
            logging.warning(f"檔案 {file.filename} 缺少必要欄位: {', '.join(missing_columns)}")
            return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={
                "success": False,
                "message": f"上傳的 Excel 檔案缺少必要欄位：{', '.join(missing_columns)}"
            })
        
        df.dropna(subset=["Name", "Data Type"], inplace=True)

        # 5. 找出數據型態不一致的欄位
        common_types = df.groupby("Name")["Data Type"].agg(lambda x: x.mode()[0] if not x.mode().empty else None).to_dict()
        df["Highlight"] = df.apply(
            lambda row: row["Data Type"] != common_types.get(row["Name"]),
            axis=1
        )
        inconsistent_names = df[df["Highlight"]]["Name"].unique()
        summary = df[df["Name"].isin(inconsistent_names)].groupby(["Name", "Data Type"]).size().reset_index(name="Count")
        logging.info(f"檔案 {file.filename} 分析完成，發現 {len(inconsistent_names)} 個名稱存在型態不一致。")

        # 6. 使用 openpyxl 處理 Excel 檔案：高亮並添加統計結果
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_font = Font(color="0000FF", underline="single")

        wb = None
        ws = None
        original_sheet_name_for_link = "Sheet1" 
        output_filename = f"result_{file.filename}" 

        if file_ext == ".xls":
            logging.info(f"處理 .xls 檔案: {file.filename}. 將創建新的 .xlsx 輸出檔案。")
            wb = openpyxl.Workbook() 
            ws = wb.active
            try:
                xls_file_handle = pd.ExcelFile(input_path)
                if xls_file_handle.sheet_names:
                    ws.title = xls_file_handle.sheet_names[0]
            except Exception:
                logging.warning("無法從 .xls 獲取原始工作表名，使用預設名稱 Sheet1")
            original_sheet_name_for_link = ws.title

            for c_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=col_name)
            for r_idx, df_row_series in df.iterrows():
                excel_row_num = r_idx + 2 
                for c_idx, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=excel_row_num, column=c_idx, value=df_row_series[col_name])
            
            for r_idx, row_data in df.iterrows():
                if row_data["Highlight"]:
                    excel_row_to_highlight = r_idx + 2
                    if excel_row_to_highlight <= ws.max_row:
                         for cell_obj in ws[excel_row_to_highlight]: 
                            cell_obj.fill = yellow_fill
            output_filename = f"result_{file_base_name}.xlsx" 

        elif file_ext in [".xlsx", ".xlsm"]:
            if file_ext == ".xlsm":
                logging.info(f"處理 .xlsm 檔案: {file.filename}. 將嘗試保留 VBA 宏。")
                wb = load_workbook(input_path, keep_vba=True)
            else: 
                logging.info(f"處理 .xlsx 檔案: {file.filename}.")
                wb = load_workbook(input_path)
            
            ws = wb[wb.sheetnames[0]] 
            original_sheet_name_for_link = ws.title

            for r_idx, row_data in df.iterrows():
                if row_data["Highlight"]:
                    excel_row = r_idx + 2 
                    if excel_row <= ws.max_row:
                        for cell_obj in ws[excel_row]:
                            cell_obj.fill = yellow_fill
        
        if "統計結果" in wb.sheetnames:
            wb.remove(wb["統計結果"])
            logging.info("已移除舊的 '統計結果' 工作表。")

        summary_ws = wb.create_sheet("統計結果")
        summary_ws.append(["Name", "Data Type", "Count"]) 

        for idx, summary_row_data in summary.iterrows():
            summary_ws.append(summary_row_data.tolist())
        logging.info("已將統計結果寫入 '統計結果' 工作表。")
        
        name_col_idx = -1
        for col_idx_loop, cell_in_header in enumerate(ws[1], start=1):
            if cell_in_header.value == "Name": 
                name_col_idx = col_idx_loop
                break
        
        if name_col_idx == -1:
            logging.warning(f"在工作表 '{original_sheet_name_for_link}' 中未找到 'Name' 欄位！無法生成超連結。")
        else:
            q_sheet_name_for_link = quote_sheetname(original_sheet_name_for_link)
            for i, row_obj_tuple in enumerate(summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, values_only=False), start=2):
                name_cell_in_summary = row_obj_tuple[0]
                field_name_in_summary = name_cell_in_summary.value

                if field_name_in_summary is None: continue

                found_row = -1
                for j in range(2, ws.max_row + 1): 
                    original_cell_value = ws.cell(row=j, column=name_col_idx).value
                    if original_cell_value is not None:
                        if str(original_cell_value).strip().lower() == str(field_name_in_summary).strip().lower():
                            found_row = j
                            break
                
                if found_row != -1:
                    target_col_letter = get_column_letter(name_col_idx)
                    internal_link_target = f"#{q_sheet_name_for_link}!{target_col_letter}{found_row}"
                    escaped_display_text = str(field_name_in_summary).replace('"', '""')
                    name_cell_in_summary.value = f'=HYPERLINK("{internal_link_target}", "{escaped_display_text}")'
                    name_cell_in_summary.font = blue_font
            logging.info("已為統計結果添加超連結。")

        # 7. 保存處理後的檔案
        output_path = os.path.join(TEMP_OUTPUT_DIR, output_filename)
        wb.save(output_path)
        logging.info(f"處理後檔案已保存到: {output_path}")

        # 8. 返回 JSON 成功響應
        return JSONResponse(status_code=status.HTTP_200_OK, content={
            "success": True,
            "message": "檔案處理完成！",
            "download_filename": output_filename
        })

    except ValueError as ve: 
        logging.error(f"處理錯誤 (ValueError): {ve}")
        return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={
            "success": False,
            "message": str(ve)
        })
    except Exception as e: 
        logging.critical(f"處理檔案 {file.filename if file else 'N/A'} 時發生未預期錯誤: {e}", exc_info=True)
        return JSONResponse(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, content={
            "success": False,
            "message": f"伺服器內部錯誤，請聯繫管理員或稍後再試。錯誤訊息: {str(e)}"
        })
    finally:
        if temp_input_dir and os.path.exists(temp_input_dir):
            try:
                shutil.rmtree(temp_input_dir)
                logging.info(f"已清理臨時輸入目錄: {temp_input_dir}")
            except OSError as e:
                logging.error(f"清理臨時輸入目錄失敗 {temp_input_dir}: {e}")

@app.get("/download/{filename}", response_class=FileResponse)
async def download_file(filename: str):
    """
    提供處理後的檔案下載。
    """
    file_path = os.path.join(TEMP_OUTPUT_DIR, filename)
    if not os.path.exists(file_path) or \
       not os.path.isfile(file_path) or \
       not os.path.abspath(file_path).startswith(os.path.abspath(TEMP_OUTPUT_DIR)):
        logging.warning(f"請求下載的檔案不存在或無效: {filename} (路徑: {file_path})")
        raise HTTPException(status_code=404, detail="檔案未找到或已過期。")

    logging.info(f"提供檔案下載: {filename}")
    return FileResponse(file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    import uvicorn
    if not os.path.exists("frontend"):
        os.makedirs("frontend")
        logging.warning("frontend 目錄不存在，已自動創建。請將 index.html 放入其中。")

    print("\n\n--- 伺服器啟動中 ---")
    print(f"請在瀏覽器中開啟: http://127.0.0.1:8000/")
    print("您可以透過拖曳或點擊來上傳 Excel 檔案。")
    print("支援格式: .xlsx, .xls, .xlsm")
    print("提醒: 處理 .xls 檔案時，Pandas 需要 'xlrd' 套件 (pip install xlrd)。")
    print("---------------------\n\n")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)